"""Export all tracker and config data to a single .xlsx backup file."""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import db
import config_db


def _fmt_date(iso_str: str | None) -> str:
    """Convert an ISO timestamp to a readable date/time string."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%Y-%m-%d %I:%M %p")
    except (ValueError, TypeError):
        return str(iso_str)


def _yn(val) -> str:
    """Convert a truthy/falsy DB value to Yes/No."""
    return "Yes" if val else "No"


def _style_sheet(ws, long_text_cols: set[int] | None = None):
    """Apply formatting: bold frozen header, auto-fit short columns, wrap long ones."""
    long_text_cols = long_text_cols or set()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    for cell in ws[1]:
        cell.font = bold

    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        if col_idx in long_text_cols:
            ws.column_dimensions[get_column_letter(col_idx)].width = 60
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = wrap
        else:
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = top
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, min(len(val), 40))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 10)


def export_backup(dest: str | Path) -> Path:
    """Write a full backup to *dest* and return the Path written."""
    dest = Path(dest)
    wb = Workbook()

    # ── Sheet 1: Openings ─────────────────────────────────────────────────
    ws_openings = wb.active
    ws_openings.title = "Openings"

    # Build a lookup: job_opening_id -> comma-separated people names
    people_by_job: dict[int, list[str]] = {}
    with db.get_connection() as conn:
        for row in conn.execute(
            "SELECT jop.job_opening_id, p.name "
            "FROM job_opening_people jop JOIN people p ON p.id = jop.person_id "
            "ORDER BY jop.job_opening_id, p.name"
        ):
            people_by_job.setdefault(row["job_opening_id"], []).append(row["name"] or "")

    headers_openings = [
        "ID", "Company", "Job Title", "Resume", "Job Description",
        "Application URL", "Archived", "People", "Date Created", "Last Updated",
    ]
    ws_openings.append(headers_openings)

    with db.get_connection() as conn:
        jobs = conn.execute("SELECT * FROM job_openings ORDER BY id").fetchall()

    for j in jobs:
        ws_openings.append([
            j["id"],
            j["customer_name"],
            j["job_title"],
            j["teal_import_raw"] or "",
            j["job_description_contents"] or "",
            j["application_url"] or "",
            _yn(j["archived"]),
            ", ".join(people_by_job.get(j["id"], [])),
            _fmt_date(j["created_at"]),
            _fmt_date(j["updated_at"]),
        ])

    # long-text columns: Resume (4), Job Description (5)
    _style_sheet(ws_openings, long_text_cols={4, 5})

    # ── Sheet 2: People ───────────────────────────────────────────────────
    ws_people = wb.create_sheet("People")

    # Build lookup: person_id -> comma-separated "Company - Job Title"
    openings_by_person: dict[int, list[str]] = {}
    with db.get_connection() as conn:
        for row in conn.execute(
            "SELECT jop.person_id, jo.customer_name, jo.job_title "
            "FROM job_opening_people jop JOIN job_openings jo ON jo.id = jop.job_opening_id "
            "ORDER BY jop.person_id, jo.customer_name"
        ):
            label = " - ".join(filter(None, [row["customer_name"], row["job_title"]]))
            openings_by_person.setdefault(row["person_id"], []).append(label)

    headers_people = [
        "ID", "Name", "Title", "Details", "Linked Openings",
        "Date Created", "Last Updated",
    ]
    ws_people.append(headers_people)

    with db.get_connection() as conn:
        people = conn.execute("SELECT * FROM people ORDER BY id").fetchall()

    for p in people:
        ws_people.append([
            p["id"],
            p["name"] or "",
            p["title"] or "",
            p["details"] or "",
            ", ".join(openings_by_person.get(p["id"], [])),
            _fmt_date(p["created_at"]),
            _fmt_date(p["updated_at"]),
        ])

    # long-text columns: Details (4)
    _style_sheet(ws_people, long_text_cols={4})

    # ── Sheet 3: Prompt Actions ───────────────────────────────────────────
    ws_actions = wb.create_sheet("Prompt Actions")

    headers_actions = [
        "ID", "Name", "Instructions", "Use Company", "Use Job Description",
        "Use Resume", "Use Person", "Person Required", "Sort Order",
        "Date Created", "Last Updated",
    ]
    ws_actions.append(headers_actions)

    actions = config_db.list_actions()
    for a in actions:
        ws_actions.append([
            a["id"],
            a["name"],
            a["instructions"] or "",
            _yn(a["use_company"]),
            _yn(a["use_job_desc"]),
            _yn(a["use_resume"]),
            _yn(a["use_person"]),
            _yn(a["person_required"]),
            a["sort_order"],
            _fmt_date(a["created_at"]),
            _fmt_date(a["updated_at"]),
        ])

    # long-text columns: Instructions (3)
    _style_sheet(ws_actions, long_text_cols={3})

    wb.save(dest)
    return dest
